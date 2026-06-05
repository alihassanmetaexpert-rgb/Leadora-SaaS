"""
Aesthetic Clinic Lead Scraper - Desktop App
============================================
A simple GUI app with search bar and one-click button.

Requirements:
    pip install requests beautifulsoup4 openpyxl selenium webdriver-manager

Usage:
    python scraper_app.py
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import time
import re
import ssl
import os
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

ssl._create_default_https_context = ssl._create_unverified_context
os.environ['WDM_SSL_VERIFY'] = '0'
os.environ['PYTHONHTTPSVERIFY'] = '0'

SAVE_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "LeadResults")
os.makedirs(SAVE_FOLDER, exist_ok=True)

SCROLL_PAUSE = 2.5
DETAIL_WAIT  = 3.5
EMAIL_REGEX  = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
SKIP_DOMAINS = ["sentry.io","wix.com","google.com","facebook.com","example.com","domain.com"]
CONTACT_PATHS = ["contact","contact-us","about","about-us","reach-us","get-in-touch"]
REQ_HEADERS  = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36"}


# ── Scraper Logic ─────────────────────────────────────────────────────────────

def build_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--window-size=1400,900")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"})
    return driver


def scroll_and_collect(driver, max_leads, log):
    wait = WebDriverWait(driver, 12)
    try:
        panel = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@role='feed']")))
    except TimeoutException:
        log("Could not find results panel.")
        return []

    links, seen, no_new = [], set(), 0
    log("Scrolling through results...")
    while len(links) < max_leads and no_new < 6:
        items = driver.find_elements(By.XPATH, "//div[@role='feed']//a[contains(@href,'maps/place')]")
        before = len(links)
        for item in items:
            href = item.get_attribute("href")
            if href and href not in seen:
                seen.add(href)
                links.append(item)
        no_new = no_new + 1 if len(links) == before else 0
        driver.execute_script("arguments[0].scrollTop += 800", panel)
        time.sleep(SCROLL_PAUSE)
    log(f"Found {len(links)} listings.")
    return links[:max_leads]


def scrape_detail(driver, link_el):
    data = {k:"" for k in ["name","category","address","city","phone","website","email","rating","maps_url"]}
    try:
        data["maps_url"] = link_el.get_attribute("href") or ""
        link_el.click()
        time.sleep(DETAIL_WAIT)
        wait = WebDriverWait(driver, 8)
        try:
            data["name"] = wait.until(EC.presence_of_element_located(
                (By.XPATH,"//h1[contains(@class,'DUwDvf')]"))).text.strip()
        except TimeoutException:
            pass
        try:
            data["category"] = driver.find_element(By.XPATH,"//button[@jsaction and contains(@class,'DkEaL')]").text.strip()
        except NoSuchElementException:
            pass
        try:
            data["rating"] = driver.find_element(By.XPATH,"//div[@class='F7nice ']//span[@aria-hidden='true']").text.strip()
        except NoSuchElementException:
            pass
        try:
            raw = driver.find_element(By.XPATH,"//button[@data-item-id='address']//div[contains(@class,'Io6YTe')]").text.strip()
            data["address"] = raw
            parts = [p.strip() for p in raw.split(",")]
            data["city"] = parts[-2] if len(parts)>=2 else (parts[-1] if parts else "")
        except NoSuchElementException:
            pass
        try:
            data["phone"] = driver.find_element(By.XPATH,"//button[starts-with(@data-item-id,'phone')]//div[contains(@class,'Io6YTe')]").text.strip()
        except NoSuchElementException:
            pass
        try:
            data["website"] = driver.find_element(By.XPATH,"//a[@data-item-id='authority']//div[contains(@class,'Io6YTe')]").text.strip()
        except NoSuchElementException:
            pass
    except Exception as e:
        pass
    return data


# ── Email Finder ──────────────────────────────────────────────────────────────

def is_valid_email(email):
    email = email.lower()
    for skip in SKIP_DOMAINS:
        if skip in email:
            return False
    if email.endswith((".png",".jpg",".jpeg",".gif",".svg",".css",".js")):
        return False
    if "noreply" in email or "no-reply" in email:
        return False
    return True


def fetch_page(url):
    try:
        r = requests.get(url, headers=REQ_HEADERS, timeout=10, verify=False)
        if r.status_code == 200:
            return r.text
    except:
        pass
    return None


def find_email(base_url):
    if not base_url:
        return ""
    if not base_url.startswith("http"):
        base_url = "https://" + base_url
    all_emails = []
    html = fetch_page(base_url)
    if html:
        all_emails += [e for e in EMAIL_REGEX.findall(html) if is_valid_email(e)]
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            if a["href"].startswith("mailto:"):
                e = a["href"].replace("mailto:","").split("?")[0].strip()
                if e and is_valid_email(e):
                    all_emails.append(e)
    if not all_emails:
        base = base_url.rstrip("/")
        for path in CONTACT_PATHS:
            html = fetch_page(f"{base}/{path}")
            if html:
                emails = [e for e in EMAIL_REGEX.findall(html) if is_valid_email(e)]
                if emails:
                    all_emails += emails
                    break
            time.sleep(0.3)
    seen = []
    for e in all_emails:
        if e not in seen:
            seen.append(e)
    return seen[0] if seen else ""


# ── Excel Export ──────────────────────────────────────────────────────────────

def export_excel(leads, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"Aesthetic Clinic Leads — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(name="Arial", bold=True, size=13, color="1A73E8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    HEADERS    = ["#","Name","Category","City","Address","Phone","Email","Website","Rating","Maps URL"]
    COL_WIDTHS = [5,  35,    22,        18,    40,       18,     32,     32,       8,       50]

    def thin():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    hfill = PatternFill("solid", start_color="1A73E8")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for ci,(h,w) in enumerate(zip(HEADERS,COL_WIDTHS),1):
        cell = ws.cell(row=2,column=ci,value=h)
        cell.font,cell.fill,cell.border = hfont,hfill,thin()
        cell.alignment = Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    odd  = PatternFill("solid", start_color="F0F4FF")
    even = PatternFill("solid", start_color="FFFFFF")
    bfont  = Font(name="Arial", size=10)
    left   = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    center = Alignment(horizontal="center",vertical="center")

    for ri,lead in enumerate(leads,3):
        fill = odd if ri%2==1 else even
        row  = [ri-2,lead.get("name",""),lead.get("category",""),lead.get("city",""),
                lead.get("address",""),lead.get("phone",""),lead.get("email",""),
                lead.get("website",""),lead.get("rating",""),lead.get("maps_url","")]
        ws.row_dimensions[ri].height = 17
        for ci,val in enumerate(row,1):
            cell = ws.cell(row=ri,column=ci,value=val)
            cell.font,cell.fill,cell.border = bfont,fill,thin()
            cell.alignment = center if ci in (1,9) else left

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{len(leads)+2}"
    wb.save(path)


# ── GUI App ───────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Aesthetic Clinic Lead Scraper")
        self.geometry("700x600")
        self.resizable(False, False)
        self.configure(bg="#f0f4ff")
        self._running = False
        self._build_ui()

    def _build_ui(self):
        # Header
        tk.Label(self, text="🏥 Aesthetic Clinic Lead Scraper",
                 font=("Arial", 18, "bold"), bg="#1A73E8", fg="white",
                 pady=14).pack(fill="x")

        # Search frame
        frame = tk.Frame(self, bg="#f0f4ff", pady=16)
        frame.pack(fill="x", padx=30)

        tk.Label(frame, text="Search Query:", font=("Arial",11,"bold"),
                 bg="#f0f4ff").grid(row=0,column=0,sticky="w",pady=4)
        self.query_var = tk.StringVar(value="aesthetic clinic")
        tk.Entry(frame, textvariable=self.query_var, font=("Arial",12),
                 width=40, relief="solid").grid(row=0,column=1,padx=10,pady=4)

        tk.Label(frame, text="City / Location:", font=("Arial",11,"bold"),
                 bg="#f0f4ff").grid(row=1,column=0,sticky="w",pady=4)
        self.city_var = tk.StringVar(value="New York, USA")
        self.city_entry = tk.Entry(frame, textvariable=self.city_var, font=("Arial",12),
                                   width=40, relief="solid")
        self.city_entry.grid(row=1,column=1,padx=10,pady=4)

        tk.Label(frame, text="Max Leads:", font=("Arial",11,"bold"),
                 bg="#f0f4ff").grid(row=2,column=0,sticky="w",pady=4)
        self.limit_var = tk.IntVar(value=40)
        tk.Spinbox(frame, from_=10, to=200, textvariable=self.limit_var,
                   font=("Arial",12), width=10, relief="solid").grid(row=2,column=1,sticky="w",padx=10,pady=4)

        self.email_var = tk.BooleanVar(value=True)
        tk.Checkbutton(frame, text="Also find emails from websites",
                       variable=self.email_var, font=("Arial",11),
                       bg="#f0f4ff").grid(row=3,column=1,sticky="w",padx=10,pady=4)

        # Button
        self.btn = tk.Button(self, text="▶  START SCRAPING",
                             font=("Arial",14,"bold"), bg="#1A73E8", fg="white",
                             activebackground="#1557b0", activeforeground="white",
                             pady=10, relief="flat", cursor="hand2",
                             command=self._start)
        self.btn.pack(fill="x", padx=30, pady=8)

        self.stop_btn = tk.Button(self, text="⏹  STOP",
                                  font=("Arial",11), bg="#e53935", fg="white",
                                  relief="flat", cursor="hand2", state="disabled",
                                  command=self._stop)
        self.stop_btn.pack(fill="x", padx=30, pady=2)

        # Progress
        self.progress = ttk.Progressbar(self, mode="indeterminate")
        self.progress.pack(fill="x", padx=30, pady=8)

        # Log
        tk.Label(self, text="Live Log:", font=("Arial",10,"bold"),
                 bg="#f0f4ff").pack(anchor="w", padx=30)
        self.log_box = scrolledtext.ScrolledText(self, height=14, font=("Courier",9),
                                                  bg="#1e1e1e", fg="#00ff88",
                                                  relief="flat", state="disabled")
        self.log_box.pack(fill="both", padx=30, pady=4)

        # Status bar
        self.status_var = tk.StringVar(value="Ready — Enter city and click Start")
        tk.Label(self, textvariable=self.status_var, font=("Arial",10),
                 bg="#1A73E8", fg="white", pady=5).pack(fill="x", side="bottom")

    def log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self.update_idletasks()

    def _start(self):
        if self._running:
            return
        self._running = True
        self.btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.progress.start(10)
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0","end")
        self.log_box.configure(state="disabled")
        threading.Thread(target=self._run, daemon=True).start()

    def _stop(self):
        self._running = False
        self.log("⏹ Stopped by user.")
        self._reset_ui()

    def _reset_ui(self):
        self.btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        self.progress.stop()

    def _run(self):
        query  = self.query_var.get().strip()
        city   = self.city_var.get().strip()
        limit  = self.limit_var.get()
        do_email = self.email_var.get()

        if not city:
            messagebox.showerror("Error", "Please enter a city or location!")
            self._reset_ui()
            return

        safe_city = city.replace(" ","_").replace(",","")
        filename  = f"leads_{safe_city}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath  = os.path.join(SAVE_FOLDER, filename)

        self.status_var.set(f"Scraping {query} in {city}...")
        self.log(f"Starting scraper...")
        self.log(f"Query  : {query}")
        self.log(f"City   : {city}")
        self.log(f"Limit  : {limit}")
        self.log(f"Output : {filepath}")
        self.log("-" * 50)

        driver = None
        leads  = []

        try:
            self.log("Opening Chrome...")
            driver = build_driver()
            full_query = f"{query} {city}"
            url = f"https://www.google.com/maps/search/{full_query.replace(' ','+')}"
            self.log(f"Searching: {full_query}")
            driver.get(url)
            time.sleep(4)

            try:
                btn = driver.find_element(By.XPATH,"//button[contains(.,'Accept') or contains(.,'Agree')]")
                btn.click()
                time.sleep(1)
            except NoSuchElementException:
                pass

            links = scroll_and_collect(driver, limit, self.log)

            for i, link in enumerate(links, 1):
                if not self._running:
                    break
                self.log(f"[{i}/{len(links)}] Scraping...", )
                d = scrape_detail(driver, link)
                if d["name"]:
                    leads.append(d)
                    self.log(f"   ✓ {d['name']} | {d['phone']}")
                else:
                    self.log(f"   — skipped")
                self.status_var.set(f"Collected {len(leads)} leads...")
                time.sleep(1)

        except Exception as e:
            self.log(f"Error: {e}")
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass

        if not leads:
            self.log("No leads found. Try a different city or query.")
            self._reset_ui()
            return

        # Find emails
        if do_email and self._running:
            self.log("-" * 50)
            self.log("Finding emails from websites...")
            for i, lead in enumerate(leads):
                if not self._running:
                    break
                site = lead.get("website","")
                if site:
                    self.log(f"   Checking: {lead['name']}...", )
                    email = find_email(site)
                    lead["email"] = email
                    if email:
                        self.log(f"   ✉ {email}")
                    else:
                        self.log(f"   — no email")
                    self.status_var.set(f"Finding emails... {i+1}/{len(leads)}")
                    time.sleep(1)

        # Save Excel
        export_excel(leads, filepath)
        self.log("-" * 50)
        self.log(f"✅ DONE! Saved {len(leads)} leads")
        self.log(f"📁 File: {filepath}")
        self.log(f"📞 Phones : {sum(1 for l in leads if l.get('phone'))}")
        self.log(f"📧 Emails : {sum(1 for l in leads if l.get('email'))}")
        self.log(f"🌐 Sites  : {sum(1 for l in leads if l.get('website'))}")
        self.status_var.set(f"✅ Done! {len(leads)} leads saved to Desktop/LeadResults")

        messagebox.showinfo("Done!", f"✅ {len(leads)} leads saved!\n\nFile: {filepath}")
        self._reset_ui()


if __name__ == "__main__":
    app = App()
    app.mainloop()
 
