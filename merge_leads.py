"""
Lead Files Merger
=================
Combines all Excel lead files from Desktop/LeadResults
into one single master Excel file with no duplicates.

Usage:
    python merge_leads.py
"""

import os
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading


LEAD_FOLDER  = os.path.join(os.path.expanduser("~"), "Desktop", "LeadResults")
SAVE_FOLDER  = LEAD_FOLDER
HEADERS      = ["#","Name","Category","City","Address","Phone","Email","Website","Rating","Maps URL"]
COL_WIDTHS   = [5,  35,    22,        18,    40,       18,     32,     32,       8,       50]


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def read_leads_from_file(filepath):
    """Read all lead rows from an Excel file."""
    leads = []
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        # Find header row (row 2)
        col_map = {}
        for cell in ws[2]:
            if cell.value:
                col_map[str(cell.value).strip()] = cell.column - 1

        if not col_map:
            return leads

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            lead = {}
            for key in ["Name","Category","City","Address","Phone","Email","Website","Rating","Maps URL"]:
                idx = col_map.get(key)
                lead[key] = str(row[idx]).strip() if idx is not None and row[idx] else ""
            if lead.get("Name") and lead["Name"] != "None":
                leads.append(lead)
    except Exception as e:
        pass
    return leads


def export_master(leads, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Leads"

    # Title
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"Master Lead Database — {datetime.now().strftime('%Y-%m-%d %H:%M')} — {len(leads)} Total Leads"
    c.font = Font(name="Arial", bold=True, size=13, color="1A73E8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    hfill = PatternFill("solid", start_color="1A73E8")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for ci,(h,w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font, cell.fill, cell.border = hfont, hfill, thin()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    # Data rows
    odd  = PatternFill("solid", start_color="F0F4FF")
    even = PatternFill("solid", start_color="FFFFFF")
    bfont  = Font(name="Arial", size=10)
    left   = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    center = Alignment(horizontal="center",vertical="center")

    for ri, lead in enumerate(leads, 3):
        fill = odd if ri % 2 == 1 else even
        row  = [ri-2,
                lead.get("Name",""), lead.get("Category",""), lead.get("City",""),
                lead.get("Address",""), lead.get("Phone",""), lead.get("Email",""),
                lead.get("Website",""), lead.get("Rating",""), lead.get("Maps URL","")]
        ws.row_dimensions[ri].height = 17
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font, cell.fill, cell.border = bfont, fill, thin()
            cell.alignment = center if ci in (1,9) else left

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{len(leads)+2}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20
    for r,(label,val) in enumerate([
        ("Total Leads",    len(leads)),
        ("With Phone",     sum(1 for l in leads if l.get("Phone"))),
        ("With Email",     sum(1 for l in leads if l.get("Email"))),
        ("With Website",   sum(1 for l in leads if l.get("Website"))),
        ("Generated",      datetime.now().strftime("%Y-%m-%d %H:%M")),
    ], 1):
        ws2.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
        ws2.cell(row=r, column=2, value=val).font   = Font(name="Arial", size=11)

    wb.save(filepath)


# ── GUI ───────────────────────────────────────────────────────────────────────

class MergerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Lead Files Merger")
        self.resizable(True, True)
        self.minsize(650, 550)
        self.state("zoomed")
        self.configure(bg="#f0f4ff")
        self._build_ui()
        self._scan_files()

    def _build_ui(self):
        # Header
        tk.Label(self, text="📊  Lead Files Merger",
                 font=("Arial", 20, "bold"), bg="#2e7d32", fg="white",
                 pady=16).pack(fill="x")

        tk.Label(self, text="Combines all your Excel lead files into ONE master sheet with no duplicates",
                 font=("Arial",11), bg="#f0f4ff", fg="#444").pack(pady=6)

        # Files list frame
        files_frame = tk.LabelFrame(self, text="  Excel Files Found in Desktop/LeadResults  ",
                                     font=("Arial",11,"bold"), bg="#f0f4ff", pady=8, padx=10)
        files_frame.pack(fill="x", padx=40, pady=8)

        # Scrollable file list with checkboxes
        self.file_vars = {}
        self.files_inner = tk.Frame(files_frame, bg="#f0f4ff")
        self.files_inner.pack(fill="x")

        # Buttons row
        btn_frame = tk.Frame(self, bg="#f0f4ff")
        btn_frame.pack(fill="x", padx=40, pady=6)
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)
        btn_frame.columnconfigure(2, weight=1)

        tk.Button(btn_frame, text="✅ Select All", font=("Arial",11),
                  bg="#0f9d58", fg="white", relief="flat", cursor="hand2",
                  command=self._select_all).grid(row=0, column=0, sticky="ew", padx=4)

        tk.Button(btn_frame, text="☐ Deselect All", font=("Arial",11),
                  bg="#888", fg="white", relief="flat", cursor="hand2",
                  command=self._deselect_all).grid(row=0, column=1, sticky="ew", padx=4)

        tk.Button(btn_frame, text="🔄 Refresh List", font=("Arial",11),
                  bg="#1A73E8", fg="white", relief="flat", cursor="hand2",
                  command=self._scan_files).grid(row=0, column=2, sticky="ew", padx=4)

        # Merge button
        self.merge_btn = tk.Button(self, text="🔗  MERGE ALL SELECTED FILES",
                                    font=("Arial",15,"bold"), bg="#2e7d32", fg="white",
                                    activebackground="#1b5e20", pady=12, relief="flat",
                                    cursor="hand2", command=self._start_merge)
        self.merge_btn.pack(fill="x", padx=40, pady=8)

        # Progress
        self.progress = ttk.Progressbar(self, mode="indeterminate")
        self.progress.pack(fill="x", padx=40, pady=4)

        # Stats
        stats_frame = tk.Frame(self, bg="#e8f0fe", pady=8)
        stats_frame.pack(fill="x", padx=40, pady=2)
        stats_frame.columnconfigure((0,1,2,3), weight=1)
        self.stat_files  = tk.StringVar(value="Files: 0")
        self.stat_total  = tk.StringVar(value="Total Rows: 0")
        self.stat_unique = tk.StringVar(value="Unique: 0")
        self.stat_dupes  = tk.StringVar(value="Duplicates Removed: 0")
        for i,(var,color) in enumerate(zip(
            [self.stat_files, self.stat_total, self.stat_unique, self.stat_dupes],
            ["#1A73E8","#0f9d58","#f4511e","#ab47bc"])):
            tk.Label(stats_frame, textvariable=var, font=("Arial",11,"bold"),
                     fg=color, bg="#e8f0fe").grid(row=0, column=i, padx=8)

        # Log
        tk.Label(self, text="Log:", font=("Arial",11,"bold"),
                 bg="#f0f4ff").pack(anchor="w", padx=40)
        self.log_box = scrolledtext.ScrolledText(self, font=("Courier",10),
                                                  bg="#1e1e1e", fg="#00ff88",
                                                  relief="flat", state="disabled", height=10)
        self.log_box.pack(fill="both", expand=True, padx=40, pady=(4,0))

        # Status
        self.status_var = tk.StringVar(value="Ready — Select files and click Merge")
        tk.Label(self, textvariable=self.status_var, font=("Arial",11),
                 bg="#2e7d32", fg="white", pady=7).pack(fill="x", side="bottom")

    def _scan_files(self):
        for w in self.files_inner.winfo_children():
            w.destroy()
        self.file_vars.clear()

        if not os.path.exists(LEAD_FOLDER):
            tk.Label(self.files_inner, text="LeadResults folder not found on Desktop.",
                     font=("Arial",10), bg="#f0f4ff", fg="red").pack()
            return

        files = [f for f in os.listdir(LEAD_FOLDER)
                 if f.endswith(".xlsx") and not f.startswith("MASTER_")]

        if not files:
            tk.Label(self.files_inner, text="No Excel files found in Desktop/LeadResults.",
                     font=("Arial",10), bg="#f0f4ff", fg="#888").pack()
            return

        for f in sorted(files):
            var = tk.BooleanVar(value=True)
            self.file_vars[f] = var
            size = os.path.getsize(os.path.join(LEAD_FOLDER, f))
            size_str = f"{size//1024} KB" if size > 1024 else f"{size} B"
            tk.Checkbutton(self.files_inner, text=f"  {f}  ({size_str})",
                           variable=var, font=("Arial",10),
                           bg="#f0f4ff", activebackground="#f0f4ff").pack(anchor="w", pady=1)

        self.stat_files.set(f"Files: {len(files)}")
        self.log(f"Found {len(files)} Excel files in LeadResults folder.")

    def _select_all(self):
        for var in self.file_vars.values():
            var.set(True)

    def _deselect_all(self):
        for var in self.file_vars.values():
            var.set(False)

    def log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self.update_idletasks()

    def _start_merge(self):
        selected = [f for f,v in self.file_vars.items() if v.get()]
        if not selected:
            messagebox.showerror("Error", "Please select at least one file!")
            return
        self.merge_btn.configure(state="disabled")
        self.progress.start(10)
        threading.Thread(target=self._merge, args=(selected,), daemon=True).start()

    def _merge(self, selected_files):
        self.log(f"\nMerging {len(selected_files)} files...")
        self.status_var.set("Merging files...")

        all_leads  = []
        total_rows = 0
        seen_names  = set()
        seen_phones = set()

        for filename in selected_files:
            filepath = os.path.join(LEAD_FOLDER, filename)
            self.log(f"  Reading: {filename}")
            leads = read_leads_from_file(filepath)
            self.log(f"   → {len(leads)} rows found")
            total_rows += len(leads)

            for lead in leads:
                name  = lead.get("Name","").strip().lower()
                phone = lead.get("Phone","").strip()

                if not name or name == "none":
                    continue
                if name in seen_names:
                    continue
                if phone and phone in seen_phones:
                    continue

                seen_names.add(name)
                if phone:
                    seen_phones.add(phone)
                all_leads.append(lead)

        dupes = total_rows - len(all_leads)

        self.stat_total.set(f"Total Rows: {total_rows}")
        self.stat_unique.set(f"Unique: {len(all_leads)}")
        self.stat_dupes.set(f"Duplicates Removed: {dupes}")

        if not all_leads:
            self.log("No leads found in selected files.")
            self.merge_btn.configure(state="normal")
            self.progress.stop()
            return

        # Save master file
        filename = f"MASTER_leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath = os.path.join(SAVE_FOLDER, filename)
        self.log(f"\nSaving master file...")
        export_master(all_leads, filepath)

        self.log(f"\n✅ DONE!")
        self.log(f"📊 Total rows read   : {total_rows}")
        self.log(f"🗑  Duplicates removed: {dupes}")
        self.log(f"✅ Unique leads saved : {len(all_leads)}")
        self.log(f"📁 Saved to          : {filepath}")
        self.status_var.set(f"✅ Done! {len(all_leads)} unique leads → {filename}")

        self.merge_btn.configure(state="normal")
        self.progress.stop()

        messagebox.showinfo("✅ Merge Complete!",
            f"Master file created!\n\n"
            f"📂 Files merged     : {len(selected_files)}\n"
            f"📊 Total rows read  : {total_rows}\n"
            f"🗑  Duplicates removed: {dupes}\n"
            f"✅ Unique leads     : {len(all_leads)}\n\n"
            f"📁 Saved to:\nDesktop → LeadResults → {filename}")


if __name__ == "__main__":
    app = MergerApp()
    app.mainloop()
